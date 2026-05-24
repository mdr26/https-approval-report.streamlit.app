"""
generate_report.py
Output: 5 sheets
  1. All Approval Summary  — deduplicated detail (1 row per ApprovalNum+Status)
  2. Sheet1                — pivot: distinct ApprovalNum count by CoName x Status
  3. Sheet2                — rejected reasons breakdown
  4. Sheet4                — Bupa rows only
  5. Sheet5                — Partially approved rows only
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SOURCE_COLS = [
    "CoCode", "CoName", "PatientFileNum", "PatientName", "DoctorCode",
    "DoctorName", "ServiceCode", "ServiceName", "ApprovalNum", "Quantity",
    "ApprovedStatus", "ApprovedQuantity", "ApprovalDate", "ApprovalDtlID",
    "RequestDateTime", "ApprovalSentDateandTime", "ApprovalResponceDateAndTime",
    "Time1Days", "Time1Hours", "Time1Minutes", "Time2Days", "Time2Hours",
    "Time2Minutes", "ResponseRemarks",
]

REQUIRED_COLS = ["CoCode", "CoName", "ApprovedStatus", "ApprovalNum", "ResponseRemarks"]

# Styles
HEADER_FILL  = PatternFill("solid", fgColor="1A2E44")
HEADER_FONT  = Font(bold=True, color="FFFFFF")
PIVOT_FILL   = PatternFill("solid", fgColor="BDD7EE")  # light blue like sample
PIVOT_FONT   = Font(bold=True, color="000000")
GRAND_FILL   = PatternFill("solid", fgColor="9DC3E6")
GRAND_FONT   = Font(bold=True)
THIN         = Side(style="thin")
BORDER       = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _find_sheet(xl):
    best_name, best_df, best_hits = None, None, -1
    for name in xl.sheet_names:
        try:
            df = xl.parse(name)
            df.columns = df.columns.str.strip()
            hits = sum(1 for c in REQUIRED_COLS if c in df.columns)
            if hits > best_hits:
                best_hits, best_name, best_df = hits, name, df
            if hits == len(REQUIRED_COLS):
                break
        except Exception:
            continue
    if best_df is None or best_hits == 0:
        raise ValueError(
            f"Could not find the expected columns in any sheet.\n"
            f"Sheets found: {xl.sheet_names}\n"
            f"Expected columns: {REQUIRED_COLS}"
        )
    return best_name, best_df


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 55)


def _write_cell(cell, val):
    if isinstance(val, float) and pd.isna(val):
        cell.value = None
    elif hasattr(val, 'item'):
        cell.value = val.item()
    else:
        cell.value = val
    cell.border = BORDER
    cell.alignment = Alignment(vertical="center")


def _write_summary_sheet(ws, df):
    """All Approval Summary — headers row 1, data row 2+"""
    for ci, col_name in enumerate(SOURCE_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 20

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            _write_cell(ws.cell(row=ri, column=ci), val)

    ws.freeze_panes = "A2"
    _auto_width(ws)


def _write_pivot_sheet(ws, df):
    """
    Sheet1 — Pivot matching sample exactly:
      Row 3: 'Distinct Count of ApprovalNum' | 'Column Labels'
      Row 4: 'Row Labels' | Approved | Pending | Partiaily | Rejected | Grand Total
      Row 5+: insurer rows
      Last:   Grand Total row
    """
    # Build pivot
    pivot = (
        df.groupby(["CoName", "ApprovedStatus"])["ApprovalNum"]
        .nunique().unstack(fill_value=0)
    )
    for col in ["Approved", "Pending", "Partiaily", "Rejected"]:
        if col not in pivot.columns:
            pivot[col] = 0
    # Order columns
    status_cols = [c for c in ["Approved", "Pending", "Partiaily", "Rejected"]
                   if c in pivot.columns]
    pivot = pivot[status_cols]
    pivot["Grand Total"] = pivot.sum(axis=1)
    grand = pivot.sum(axis=0).rename("Grand Total")
    pivot = pd.concat([pivot, grand.to_frame().T])
    pivot.index.name = "Row Labels"
    pivot = pivot.reset_index()

    all_cols = list(pivot.columns)  # Row Labels + status cols + Grand Total

    # Row 3 — label row
    c1 = ws.cell(row=3, column=1, value="Distinct Count of ApprovalNum")
    c1.font = Font(bold=True)
    c2 = ws.cell(row=3, column=2, value="Column Labels")
    c2.font = Font(bold=True)

    # Row 4 — headers
    for ci, col in enumerate(all_cols, 1):
        cell = ws.cell(row=4, column=ci, value=col)
        cell.font = PIVOT_FONT
        cell.fill = PIVOT_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                   vertical="center")

    # Row 5+ — data rows
    for ri, row in enumerate(pivot.itertuples(index=False), 5):
        is_grand = str(row[0]) == "Grand Total"
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci)
            cell.value = val.item() if hasattr(val, 'item') else val
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal="center" if ci > 1 else "left",
                vertical="center"
            )
            if is_grand:
                cell.font = GRAND_FONT
                cell.fill = GRAND_FILL

    _auto_width(ws)


def _write_reasons_sheet(ws, df):
    """
    Sheet2 — Rejected reasons:
      Row 1: ApprovedStatus | Rejected
      Row 3: Row Labels | Count of ResponseRemarks
      Row 4+: reasons + Grand Total
    """
    ws.cell(row=1, column=1, value="ApprovedStatus").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Rejected").font = Font(bold=True)

    ws.cell(row=3, column=1, value="Row Labels").font = Font(bold=True)
    ws.cell(row=3, column=2, value="Count of ResponseRemarks").font = Font(bold=True)

    rejected = df[df["ApprovedStatus"] == "Rejected"]
    reasons  = rejected["ResponseRemarks"].value_counts().reset_index()
    reasons.columns = ["Row Labels", "Count of ResponseRemarks"]
    grand = pd.DataFrame(
        [["Grand Total", int(reasons["Count of ResponseRemarks"].sum())]],
        columns=reasons.columns,
    )
    reasons = pd.concat([reasons, grand], ignore_index=True)

    for ri, row in enumerate(reasons.itertuples(index=False), 4):
        c1 = ws.cell(row=ri, column=1, value=row[0])
        c2 = ws.cell(row=ri, column=2, value=row[1])
        if row[0] == "Grand Total":
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
        c1.border = BORDER
        c2.border = BORDER

    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 25


def _write_detail_sheet(ws, title, df):
    """Sheet4 / Sheet5 — title row1, blank row2, headers row3, data row4+"""
    ws.cell(row=1, column=1, value=title).font = Font(bold=True)

    for ci, col_name in enumerate(SOURCE_COLS, 1):
        cell = ws.cell(row=3, column=ci, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for ri, row in enumerate(df.itertuples(index=False), 4):
        for ci, val in enumerate(row, 1):
            _write_cell(ws.cell(row=ri, column=ci), val)

    ws.freeze_panes = "A4"
    _auto_width(ws)


def generate_report(approval_file, output_path):
    # ── Load & detect sheet ───────────────────────────────────────────────
    xl = pd.ExcelFile(approval_file)
    detected_sheet, df = _find_sheet(xl)
    df.columns = df.columns.str.strip()

    for col in SOURCE_COLS:
        if col not in df.columns:
            df[col] = None
    df = df[SOURCE_COLS].copy()

    rows_before = len(df)

    # ── Deduplicate: 1 row per (ApprovalNum, ApprovedStatus) ─────────────
    df = df.drop_duplicates(subset=["ApprovalNum", "ApprovedStatus"], keep="first")
    df = df.reset_index(drop=True)

    rows_removed = rows_before - len(df)

    # ── Subsets ───────────────────────────────────────────────────────────
    bupa    = df[df["CoName"].str.contains("Bupa", case=False, na=False)].copy()
    partial = df[df["ApprovedStatus"] == "Partiaily"].copy()

    # ── Build workbook ────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    # 1. All Approval Summary
    ws_all = wb.create_sheet("All Approval Summary")
    _write_summary_sheet(ws_all, df)

    # 2. Sheet1 — Pivot
    ws1 = wb.create_sheet("Sheet1")
    _write_pivot_sheet(ws1, df)

    # 3. Sheet2 — Rejection reasons
    ws2 = wb.create_sheet("Sheet2")
    _write_reasons_sheet(ws2, df)

    # 4. Sheet4 — Bupa detail
    ws4 = wb.create_sheet("Sheet4")
    _write_detail_sheet(ws4,
        "Data returned for Distinct Count of ApprovalNum, "
        "Bupa Arabia for Cooperative Insurance.", bupa)

    # 5. Sheet5 — Partial detail
    ws5 = wb.create_sheet("Sheet5")
    _write_detail_sheet(ws5,
        "Data returned for Distinct Count of ApprovalNum, Partiaily.", partial)

    wb.save(output_path)

    return {
        "total":          len(df),
        "approved":       int((df["ApprovedStatus"] == "Approved").sum()),
        "partial":        int((df["ApprovedStatus"] == "Partiaily").sum()),
        "pending":        int((df["ApprovedStatus"] == "Pending").sum()),
        "rejected":       int((df["ApprovedStatus"] == "Rejected").sum()),
        "rows_removed":   rows_removed,
        "detected_sheet": detected_sheet,
    }
